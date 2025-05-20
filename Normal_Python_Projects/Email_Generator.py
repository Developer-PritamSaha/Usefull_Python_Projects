import tkinter as tk
from tkinter import messagebox
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side

def database_updater(n,d,m,email,password):
    f = 1
    try:
        workbook = openpyxl.load_workbook('E-wealth_Emails.xlsx')
    except:
        workbook = openpyxl.Workbook()
        f = 0
        
    sheet = workbook.active
    if  f == 0:
        sheet['A1'] = 'Name'
        sheet['A1'].font = Font('Arial',12,bold=True)
        sheet['A1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet['A1'].border = Border(left = Side(style='medium'),right = Side(style='medium')  ,bottom=Side(style='thick'))
        sheet['B1'] = 'D.O.B'
        sheet['B1'].font = Font('Arial',12,bold=True)
        sheet['B1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet['B1'].border = Border(left = Side(style='medium'),right = Side(style='medium')  ,bottom=Side(style='thick'))
        sheet['C1'] = 'Mobile No.'
        sheet['C1'].font = Font('Arial',12,bold=True)
        sheet['C1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet['C1'].border = Border(left = Side(style='medium'),right = Side(style='medium')  ,bottom=Side(style='thick'))
        sheet['D1'] = 'Email'
        sheet['D1'].font = Font('Arial',12,bold=True)
        sheet['D1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet['D1'].border = Border(left = Side(style='medium'),right = Side(style='medium')  ,bottom=Side(style='thick'))
        sheet['E1'] = 'Password'
        sheet['E1'].font = Font('Arial',12,bold=True)
        sheet['E1'].fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        sheet['E1'].border = Border(left = Side(style='medium'),right = Side(style='medium')  ,bottom=Side(style='thick'))
        
    for row in range(1, sheet.max_row + 2):
        if sheet['A'+str(row)].value == None:
            temp = str(row)
            break

    sheet['A'+temp] =  n
    sheet['B'+temp] =  d
    sheet['C'+temp] =  m
    sheet['D'+temp] =  email
    sheet['E'+temp] =  password

    f = 1
    try:
        workbook.save('E-wealth_Emails.xlsx')
    except:
        messagebox.showerror("File not saved", "*Not able to Save the file* \n\n Please close the 'E-wealth_Emails.xlsx' file opened in other application and try again.")
        f = 0
    return f
    
def show_results(e,p,f):
    if f == 0:
        user_pass = '\n    <! Unable to generate email and password. !>\n\n # Kindly address the exception that has occurred,\t\t  and try again. #'
    else:
        user_pass = '\n※ Email >> ' + e + '\n\n※ Password >> ' + p
    
    # Display the result in the Text widget
    user_pass_widget.config(state="normal")
    user_pass_widget.delete("1.0", tk.END)  # Clear previous content
    user_pass_widget.insert(tk.END, user_pass)
    user_pass_widget.config(state="disabled")

def process():
    f_name = entry1.get()
    f_name = f_name.replace(' ','')
    s_name = entry2.get()
    surname = s_name.replace(' ','')
    dob = entry3.get()
    mobile = entry4.get()

    f = 1

    if len(f_name) < 2:
        messagebox.showerror("Name Error", "Please check the first name .")
        f = 0
    if len(s_name) < 2:
        messagebox.showerror("Name Error", "Please check the last name .")
        f = 0
    if dob.count('-') != 2 or len(dob) != 10:
        messagebox.showerror("D.O.B Error", "Please enter the D.O.B in (dd-mm-yyyy) format with proper hypens.")
        f = 0
    else:
        try:
            if dob[0] == '-' or int(dob[:2]) > 31:
                messagebox.showerror("D.O.B Error", "Please check the date inserted.")
                f = 0
            if dob[3] == '-' or int(dob[3:5]) > 12:
                messagebox.showerror("D.O.B Error", "Please check the month inserted.")
                f = 0
            current_year = datetime.now().year
            if dob[6] == '-' or int(dob[6:]) > (current_year-1):
                messagebox.showerror("D.O.B Error", "Please check the year inserted.")
                f = 0
        except:
            messagebox.showerror("D.O.B Error", "Please enter the D.O.B in (dd-mm-yyyy) format with proper hypens.")
            f = 0

    if len(mobile) == 10: 
        try:
            t = int(mobile)
        except:
            messagebox.showerror("Mobile No. Error", "Please remove any characters only digits are allowed.")
            f = 0
    else:
        messagebox.showerror("Mobile No. Error", "Please check the mobile no. inputed.")
        f = 0

    if len(s_name) > 3:
        surname = s_name[0:3]

    email_Add = f_name.lower() + '.' + surname.lower() + mobile[7:] + '@gmail.com'
    password = f_name[0:2].lower() + '#@' + dob[8:] + surname[0:2].lower()

    name = f_name + ' ' + s_name
    if f != 0:
        f = database_updater(name,dob,mobile,email_Add,password)

    show_results(email_Add,password,f)

window = tk.Tk()
window.title('Email-Generator')
window.geometry("1000x760+100+100")

label = tk.Label(window, font= ('Arial',22,'bold'),fg='blue', text='Welcome!')
label.pack(pady=(60,30))

custom_font = ("Helvetica", 16)

label1 = tk.Label(window,fg = 'gray',font=custom_font, text="First Name : ")
label1.pack(pady=(30,0))

entry1 = tk.Entry(window,fg='green',font=custom_font)
entry1.pack(pady=0)

label2 = tk.Label(window,fg = 'gray',font=custom_font, text="Last Name : ")
label2.pack(pady=(30,0))

entry2 = tk.Entry(window,fg='green',font=custom_font)
entry2.pack(pady=0)

label3 = tk.Label(window,fg = 'gray', font=custom_font, text="D.O.B (dd-mm-yyyy) :")
label3.pack(pady=(30,0))

entry3 = tk.Entry(window,fg='red',font=custom_font)
entry3.pack(pady=0)

label4 = tk.Label(window,fg = 'gray',font=custom_font, text="Mobile No. :")
label4.pack(pady=(30,0))

entry4 = tk.Entry(window,fg='red',font=custom_font)
entry4.pack(pady=0)

button = tk.Button(window,font=('Arial',12,'bold'),bg='sky blue',fg='navy blue',text="<<Generate>>", command=process)
button.pack(pady=(40,30))

user_pass_widget = tk.Text(window,font=('Arial',15), height=5, width=40, state="disabled")
user_pass_widget.pack(pady=(10,0))

window.mainloop()