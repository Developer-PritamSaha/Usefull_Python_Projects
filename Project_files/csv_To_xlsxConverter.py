# Convert any 'CSV' file to 'XLSX' file format

import csv
import openpyxl

file_path = input('\n#> Enter the file path or name of the file : ').replace('"','')
print(file_path)
t = len(file_path)
if file_path[t-4:].lower() != '.csv':
    raise Exception('\n <!> Error: File format is not supported! Please enter a ".csv" file.')


t = len(file_path)
save_name = file_path[0:t-4] + '.xlsx'


#>> This block enables you to choose a different file name for your extracted '.xlsx' file
# save_name = input('\n#> Enter the name of file to be saved : ')
# t = len(save_name)
# if save_name[t-5:] != '.xlsx':
#     save_name += '.xlsx'

csv_file = open(file_path,'r')
data = csv.reader(csv_file)

next = 1
workbook = openpyxl.Workbook()
sheet = workbook.active

def get_column_letter(n):
    '''This function generates the appropriate column letters for columns beyond 'Z'. This allows the script to handle CSV files with more than 26 columns.'''
    result = ''
    while n > 0:
        n, remainder = divmod(n - 1, 26)
        result = chr(65 + remainder) + result
    return result

for row in data:
    if next == 1:
        c = len(row)
    for i in range(1,c+1):
        t = str(next)
        sheet[get_column_letter(i) + t] = row[i - 1]
    next += 1

workbook.save(save_name)

print("\n ※ Your file has been converted successfully!\n   ⁘ '.xlsx' file has been saved in the provided '.csv' file's directory. \n")
